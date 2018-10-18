#!/usr/local/bin python3
#remember to install openpyxl
import openpyxl
from openpyxl.utils.cell import get_column_letter, column_index_from_string


def read_excel(name_of_excel):
    wb = openpyxl.load_workbook('{}'.format(name_of_excel))
    sheet = wb['Sheet1']
    number_of_rows = sheet.max_row
    number_of_columns = sheet.max_column
    last_column_letter = get_column_letter(number_of_columns)
    list_of_rows = []
    for row in sheet['A1' : '{}'.format(last_column_letter + str(number_of_rows))]:
        each_row = []
        for cell_object in row:
            each_row.append(cell_object.value)
        list_of_rows.append(each_row)
    return list_of_rows

    
